import openpyxl
import pandas as pd
import math

#Read and Convert .csv file into .xlsx
read_file = pd.read_csv (r'Path of .csv file from Matlab Mobile')
read_file.to_excel (r'Path of Temp folder\sensor.xlsx', index = None, header=True)
path = "Path of Temp folder\sensor.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row=sheet_obj.max_row
val=[]
for_sum=[]

#Loop to calculate the magnitude of acceleraion and normalize the effect of gravity
#Output here is a array of magnitude of acceleration
for i in range(2, row+1):
	ax = sheet_obj.cell(row = i, column = 2)
	ay = sheet_obj.cell(row = i, column = 3)
	az = sheet_obj.cell(row = i, column = 4)
	for_sum.append((math.sqrt((ax.value**2)
                    +(ay.value**2)+(az.value**2))))
	(val.append((math.sqrt((ax.value**2)
                    +(ay.value**2)+(az.value**2)))
                    -(sum(for_sum)/row)))

#Declaration of variables for peak detection algorithm	
vect=[0,0,0]
peak=[]
peak_ind=[]
base=[]
base_ind=[]
temp_l1=[]
temp_l2=[]
start=[]
end=[]
step_sz=[]
alpha=[]
dist=[]

#Peak detection Algorithm
for i in range(0,row-3):
    for j in range(0,3):
        vect[j]=val[i+j]
        if (vect[1]>vect[0] and vect[1]>vect[2] and vect[1]>12.78): #Threshold/Filter Window
            peak.append(vect[1])
            peak_ind.append(i)
        elif (vect[1]<vect[0] and vect[1]<vect[2]):
            base.append(vect[1])
            base_ind.append(i)
for i in range(0, len(peak_ind)):
    temp=peak_ind[i]
    for j in range(0, len(base_ind)):
        t1=temp-base_ind[j]
        if t1>0:
            temp_l1.append(t1)
    s=min(temp_l1)
    temp_l1.clear()
    v1=temp-s
    for j in range(0, len(base_ind)):
        if base_ind[j]==v1:
            start.append(base_ind[j])
        t2=base_ind[j]-temp
        if t2>0:
            temp_l2.append(t2)
    b=min(temp_l2)
    temp_l2.clear()
    v2=temp+b
    for j in range(0, len(base_ind)):
        if base_ind[j]==v2:
            end.append(base_ind[j])

#Multiplier as per step size
for i in range(0, len(start)):
    step_sz.append(end[i]-start[i])
    magav=(sum(for_sum)/row)
    if step_sz[i]<magav:
        avar=0.5
        alpha.append(avar)
    elif step_sz[i]==magav:
        avar=1
        alpha.append(avar)
    elif step_sz[i]>magav:
        avar=1.5
        alpha.append(avar)
    dist.append(alpha[i]*peak[i])
dist_cov=int(sum(dist)) #Calculate Distance
steps=int(dist_cov/0.8) #Estimate number of Steps

#Display output using Info Pop-up
from tkinter import *
import tkinter.messagebox
disp = tkinter.Tk()
tkinter.messagebox.showinfo("Info",
                            "You covered around %d m and walked approximately %d steps"
                            %(dist_cov,steps))
mainloop()
